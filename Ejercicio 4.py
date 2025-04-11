from openpyxl import load_workbook

libro = load_workbook("ventas.xlsx")

hoja = libro["Ventas"]

for fila in hoja.iter_rows(min_row=2, values_only=False):
    if fila[0].value == "Manzanas": 
        fila[2].value = 0.55      
        break

libro.save("mi_libro.xlsx")
print("El precio de las Manzanas se actualizó a 0.55 y se guardó en 'mi_libro.xlsx'.")
