from openpyxl import load_workbook

libro = load_workbook("ventas.xlsx")

hoja = libro["Ventas"]

datos_leidos = []
for fila in hoja.iter_rows(values_only=True):
    datos_leidos.append(list(fila))

print(datos_leidos)
