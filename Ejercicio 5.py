from openpyxl import load_workbook

libro = load_workbook("mi_libro.xlsx")

hoja = libro["Ventas"]

hoja.cell(row=1, column=4, value="Total")

for fila in range(2, hoja.max_row + 1):
    cantidad = hoja.cell(row=fila, column=2).value
    precio = hoja.cell(row=fila, column=3).value

    if isinstance(cantidad, (int, float)) and isinstance(precio, (int, float)):
        total = cantidad * precio
    else:
        total = None

    hoja.cell(row=fila, column=4, value=total)

libro.save("mi_libro.xlsx")
print("Se ha añadido la columna 'Total' y se han guardado los cambios en 'mi_libro.xlsx'.")
