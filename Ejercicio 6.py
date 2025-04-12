from openpyxl import load_workbook

libro = load_workbook("mi_libro.xlsx")

hoja_inventario = libro.create_sheet(title="Inventario")

inventario = [
    ["Producto", "Stock"],
    ["Manzanas", 100],
    ["Naranjas", 80]
]

for fila_idx, fila_datos in enumerate(inventario, start=1):
    for col_idx, valor in enumerate(fila_datos, start=1):
        hoja_inventario.cell(row=fila_idx, column=col_idx, value=valor)

libro.save("mi_libro.xlsx")

hoja_ventas = libro["Ventas"]
datos_ventas = {}
for fila in hoja_ventas.iter_rows(min_row=2, values_only=True):  # Omitir cabecera
    producto, cantidad, precio, total = fila
    datos_ventas[producto] = {"Cantidad": cantidad, "Precio": precio, "Total": total}

datos_inventario = {}
for fila in hoja_inventario.iter_rows(min_row=2, values_only=True):  # Omitir cabecera
    producto, stock = fila
    datos_inventario[producto] = {"Stock": stock}

print("Datos de Ventas:", datos_ventas)
print("Datos de Inventario:", datos_inventario)
